import io
import csv
import calendar
from datetime import datetime, date, timedelta
import os
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def generate_pdf_report(headers, data, field_sites=None, title="Report", date_range_str="All Time", records=None):
    """Generate a PDF report from headers and data arrays, with one page per farm."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch)
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Title'],
        fontSize=18, spaceAfter=6, textColor=colors.HexColor('#0b9e4f')
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle', parent=styles['Normal'],
        fontSize=10, textColor=colors.HexColor('#64748b'), spaceAfter=20
    )

    elements = []

    # If field_sites is somehow empty, process once with 'All Sites' behavior
    sites_to_process = field_sites if field_sites else [None]

    for index, site in enumerate(sites_to_process):
        # Header
        elements.append(Paragraph("Philippine Coconut Authority", title_style))
        elements.append(Paragraph(title, subtitle_style))

        site_label = site.name if site else "All Field Sites"
        generated = datetime.now().strftime('%B %d, %Y %I:%M %p')
        
        info_str = f"Field Site: {site_label} &nbsp;&nbsp;|&nbsp;&nbsp; Period: {date_range_str} &nbsp;&nbsp;|&nbsp;&nbsp; Generated: {generated}"
        elements.append(Paragraph(info_str, subtitle_style))
        elements.append(Spacer(1, 12))

        # Filter 'data' array for ONLY records matching this site.
        # Since 'data' is a pre-built array from views.py that doesn't explicitly store 'site_id' cleanly 
        # (usually it stores string name like 'Balilihan' in some varying column), the safest approach 
        # is to check if the row's string contains the site name. 
        # BUT the best approach given the current views.py logic is that views.py passes down the unfiltered `data` 
        # which already has string fields. 
        site_data = []
        if len(sites_to_process) > 1 and site:
            # We must filter the data rows where site mapping applies.
            # In reports/views.py the Farm Location or Field Site is usually in the second-to-last or last column.
            # We'll do a simple substring match on the row contents for the site name.
            for row in data:
                # convert row to string to search for site.name
                if site.name in str(row):
                    site_data.append(row)
        else:
            site_data = data

        table_data = [headers] + site_data

        if len(site_data) > 0:
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0b9e4f')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("No records found for this farm in this date range.", styles['Normal']))

        # Approval section (New Layout per User Request)
        elements.append(Spacer(1, 40))
        
        site_name = site.name if site else ""
        # Localize labels and signatories from FieldSite overrides
        prep_label = site.prepared_by_label if site and site.prepared_by_label else "Prepared by:"
        rev_label = site.reviewed_by_label if site and site.reviewed_by_label else "Reviewed by:"
        note_label = site.noted_by_label if site and site.noted_by_label else "Noted by:"

        ov_prep_name = site.prepared_by_name if site and site.prepared_by_name else ""
        ov_prep_title = site.prepared_by_title if site and site.prepared_by_title else ""
        ov_rev_name = site.reviewed_by_name if site and site.reviewed_by_name else ""
        ov_rev_title = site.reviewed_by_title if site and site.reviewed_by_title else ""
        ov_note_name = site.noted_by_name if site and site.noted_by_name else ""
        ov_note_title = site.noted_by_title if site and site.noted_by_title else ""

        prepared_user = None
        reviewed_user = None
        noted_user = None

        if records:
            site_records = records
            if site:
                site_records = [r for r in records if getattr(r, 'field_site', None) == site]
            for rec in reversed(list(site_records)):
                if not prepared_user and getattr(rec, 'prepared_by', None):
                    prepared_user = rec.prepared_by
                if not reviewed_user and getattr(rec, 'reviewed_by', None):
                    reviewed_user = rec.reviewed_by
                if not noted_user and getattr(rec, 'noted_by', None):
                    noted_user = rec.noted_by

        # Final Text Strings
        if ov_prep_name:
            prepared_by_str = f"<b>{ov_prep_name.upper()}</b><br/>{ov_prep_title}"
        elif prepared_user:
            title = prepared_user.profile.get_role_display() if hasattr(prepared_user, 'profile') else 'COS/Agriculturist'
            name_str = (prepared_user.get_full_name() or prepared_user.username).strip().upper()
            prepared_by_str = f"<b>{name_str}</b><br/>{title}"
        else:
            prepared_by_str = "_______________________<br/>COS/Agriculturist"

        if ov_rev_name:
            reviewed_by_str = f"<b>{ov_rev_name.upper()}</b><br/>{ov_rev_title}"
        elif reviewed_user:
            title = reviewed_user.profile.get_role_display() if hasattr(reviewed_user, 'profile') else 'Senior Agriculturist'
            name_str = (reviewed_user.get_full_name() or reviewed_user.username).strip().upper()
            reviewed_by_str = f"<b>{name_str}</b><br/>{title}"
        else:
            reviewed_by_str = "_______________________<br/>Senior Agriculturist"

        if ov_note_name:
            noted_by_str = f"<b>{ov_note_name.upper()}</b><br/>{ov_note_title}"
        elif noted_user:
            title = noted_user.profile.get_role_display() if hasattr(noted_user, 'profile') else 'PCDM/Division Chief I'
            name_str = (noted_user.get_full_name() or noted_user.username).strip().upper()
            noted_by_str = f"<b>{name_str}</b><br/>{title}"
        else:
            noted_by_str = "_______________________<br/>PCDM/Division Chief I"

        from reportlab.platypus import Image as RLImage
        def get_signature_flowable(user_obj):
            if user_obj and hasattr(user_obj, 'profile') and user_obj.profile.signature_image:
                try:
                    img_path = user_obj.profile.signature_image.path
                    if os.path.exists(img_path):
                        return RLImage(img_path, width=1.5*inch, height=0.6*inch)
                except Exception:
                    pass
            return Spacer(1, 0.6*inch)

        prep_img = get_signature_flowable(prepared_user) if not ov_prep_name else Spacer(1, 0.6*inch)
        rev_img = get_signature_flowable(reviewed_user) if not ov_rev_name else Spacer(1, 0.6*inch)
        note_img = get_signature_flowable(noted_user) if not ov_note_name else Spacer(1, 0.6*inch)

        centered_style = ParagraphStyle('Center', parent=styles['Normal'], alignment=1, fontSize=9)
        left_style = ParagraphStyle('Left', parent=styles['Normal'], alignment=0, fontSize=9)

        row1 = [
            Paragraph(f"<b>{prep_label}</b>", left_style),
            Paragraph(f"<b>{rev_label}</b>", left_style),
            Paragraph(f"<b>{note_label}</b>", left_style)
        ]
        row2 = [prep_img, rev_img, note_img]
        row3 = [
            Paragraph(prepared_by_str, centered_style),
            Paragraph(reviewed_by_str, centered_style),
            Paragraph(noted_by_str, centered_style)
        ]

        approval_data = [row1, row2, row3]
        
        approval_table = Table(approval_data, colWidths=[3.0*inch, 3.0*inch, 3.0*inch])
        approval_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
        ]))
        elements.append(approval_table)

        # Add PageBreak if not the last site
        if index < len(sites_to_process) - 1:
            elements.append(PageBreak())

    doc.build(elements)
    buffer.seek(0)
    return buffer

import os
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

def _excel_auto_width(ws):
    """Auto-size column widths with a minimum buffer for editing."""
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    # Handle multi-line text by taking the longest line
                    lines = str(cell.value).split('\n')
                    line_max = max([len(l) for l in lines])
                    if line_max > max_length:
                        max_length = line_max
            except:
                pass
        # Min width of 10 for better fit on Letter paper, max of 40
        adjusted_width = max(10, min(max_length + 2, 40))
        ws.column_dimensions[column].width = adjusted_width


def _add_header_logo(ws):
    """Add the PCA & DA logo to the top left of the worksheet."""
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'PCA&DA_Logo.png')
    if os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.height = 75
        img.width = int(img.height * (img.width / img.height)) if img.height else 75
        img.anchor = 'F1'
        ws.add_image(img)


def _add_footer(ws, start_row, site, num_cols, records=None):
    """Add the Prepared by, Reviewed by, and Noted by footer in a 3-column layout."""
    # Localize labels and signatories from FieldSite overrides
    prep_label = site.prepared_by_label if site and site.prepared_by_label else "Prepared by:"
    rev_label = site.reviewed_by_label if site and site.reviewed_by_label else "Reviewed by:"
    note_label = site.noted_by_label if site and site.noted_by_label else "Noted by:"

    ov_prep_name = site.prepared_by_name if site and site.prepared_by_name else ""
    ov_prep_title = site.prepared_by_title if site and site.prepared_by_title else ""
    ov_rev_name = site.reviewed_by_name if site and site.reviewed_by_name else ""
    ov_rev_title = site.reviewed_by_title if site and site.reviewed_by_title else ""
    ov_note_name = site.noted_by_name if site and site.noted_by_name else ""
    ov_note_title = site.noted_by_title if site and site.noted_by_title else ""

    prepared_user = None
    reviewed_user = None
    noted_user = None

    if records:
        for rec in reversed(list(records)):
            if not prepared_user and getattr(rec, 'prepared_by', None):
                prepared_user = rec.prepared_by
            if not reviewed_user and getattr(rec, 'reviewed_by', None):
                reviewed_user = rec.reviewed_by
            if not noted_user and getattr(rec, 'noted_by', None):
                noted_user = rec.noted_by

    # Final strings for Prepared By
    if ov_prep_name:
        prepared_by = ov_prep_name.upper()
        prepared_title = ov_prep_title
    elif prepared_user:
        mi = prepared_user.profile.middle_initial + ' ' if hasattr(prepared_user, 'profile') and prepared_user.profile.middle_initial else ''
        name_str = f"{prepared_user.first_name} {mi}{prepared_user.last_name}".strip().upper()
        prepared_by = name_str if name_str else prepared_user.username.upper()
        prepared_title = prepared_user.profile.get_role_display() if hasattr(prepared_user, 'profile') else 'COS/Agriculturist'
    else:
        prepared_by = '_______________________'
        prepared_title = 'COS/Agriculturist'

    # Reviewed By
    if ov_rev_name:
        reviewed_by = ov_rev_name.upper()
        reviewed_title = ov_rev_title
    elif reviewed_user:
        mi = reviewed_user.profile.middle_initial + ' ' if hasattr(reviewed_user, 'profile') and reviewed_user.profile.middle_initial else ''
        name_str = f"{reviewed_user.first_name} {mi}{reviewed_user.last_name}".strip().upper()
        reviewed_by = name_str if name_str else reviewed_user.username.upper()
        reviewed_title = reviewed_user.profile.get_role_display() if hasattr(reviewed_user, 'profile') else 'Senior Agriculturist'
    else:
        reviewed_by = '_______________________'
        reviewed_title = 'Senior Agriculturist'

    # Noted By
    if ov_note_name:
        noted_by = ov_note_name.upper()
        noted_title = ov_note_title
    elif noted_user:
        mi = noted_user.profile.middle_initial + ' ' if hasattr(noted_user, 'profile') and noted_user.profile.middle_initial else ''
        name_str = f"{noted_user.first_name} {mi}{noted_user.last_name}".strip().upper()
        noted_by = name_str if name_str else noted_user.username.upper()
        noted_title = noted_user.profile.get_role_display() if hasattr(noted_user, 'profile') else 'PCDM/Division Chief I'
    else:
        noted_by = '_______________________'
        noted_title = 'PCDM/Division Chief I'
    
    col1 = 1
    # Find a good middle column. If 21 cols, col2 = 10 or 11.
    col2 = max(2, (num_cols // 2) - 1)
    
    # Find a good right column that isn't cut off
    if num_cols >= 18:
        col3 = num_cols - 4
    elif num_cols >= 12:
        col3 = num_cols - 2
    else:
        col3 = max(3, num_cols - 1)
        
    row = start_row + 3
    
    center_font = Font(name='Calibri', size=10, bold=True)
    normal_font = Font(name='Calibri', size=10)
    center_align = Alignment(horizontal='center')
    left_align = Alignment(horizontal='left')

    # labels row
    ws.cell(row=row, column=col1, value=prep_label).font = center_font
    ws.cell(row=row, column=col1).alignment = left_align
    
    ws.cell(row=row, column=col2, value=rev_label).font = center_font
    ws.cell(row=row, column=col2).alignment = left_align

    ws.cell(row=row, column=col3, value=note_label).font = center_font
    ws.cell(row=row, column=col3).alignment = left_align
    
    row += 4
    
    # Names
    ws.cell(row=row, column=col1, value=prepared_by).font = center_font
    ws.cell(row=row, column=col1).alignment = center_align
    
    ws.cell(row=row, column=col2, value=reviewed_by).font = center_font
    ws.cell(row=row, column=col2).alignment = center_align
    
    ws.cell(row=row, column=col3, value=noted_by).font = center_font
    ws.cell(row=row, column=col3).alignment = center_align
    
    row += 1
    
    # Titles
    ws.cell(row=row, column=col1, value=prepared_title).font = normal_font
    ws.cell(row=row, column=col1).alignment = center_align
    
    ws.cell(row=row, column=col2, value=reviewed_title).font = normal_font
    ws.cell(row=row, column=col2).alignment = center_align
    
    ws.cell(row=row, column=col3, value=noted_title).font = normal_font
    ws.cell(row=row, column=col3).alignment = center_align
    
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    
    # Insert signatures
    def insert_signature(user_obj, col_idx):
        if not user_obj or not hasattr(user_obj, 'profile'): return
        if not user_obj.profile.signature_image: return
        try:
            img_path = user_obj.profile.signature_image.path
            if os.path.exists(img_path):
                img = XLImage(img_path)
                # Max dimensions (approx 150x60)
                img.width = min(img.width, 150)
                img.height = min(img.height, 60)
                
                # Dynamic Centering:
                col_letter = get_column_letter(col_idx)
                col_width_chars = ws.column_dimensions[col_letter].width or 20
                col_width_pixels = col_width_chars * 7.5 # Approx pixels per char in Calibri 11
                
                # Centering calculation
                offset_pixels = max(0, (col_width_pixels - img.width) / 2)
                offset_emu = int(offset_pixels * 9525)
                
                # anchor it above the name
                marker = AnchorMarker(col=col_idx-1, colOff=offset_emu, row=row-3, rowOff=0)
                size = XDRPositiveSize2D(int(img.width * 9525), int(img.height * 9525))
                img.anchor = OneCellAnchor(_from=marker, ext=size)
                ws.add_image(img)
        except Exception:
            pass

    insert_signature(prepared_user if not ov_prep_name else None, col1)
    insert_signature(reviewed_user if not ov_rev_name else None, col2)
    insert_signature(noted_user if not ov_note_name else None, col3)

    return row



def generate_excel_export(module, records, as_of_date=None):
    """Generate an Excel export with multiple sheets for Field Sites."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Group records by FieldSite name
    sites = {}
    for r in records:
        site_name = r.field_site.name if r.field_site else "Unknown Site"
        if site_name not in sites:
            sites[site_name] = []
        sites[site_name].append(r)

    if not sites:
        ws = wb.create_sheet("No Data")
        ws.append(["No records found in this date range."])
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # Define styles
    header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    title_font = Font(bold=True)

    # Build the "as of" date string
    if as_of_date:
        as_of_str = f'as of {as_of_date.strftime("%B %d, %Y")}'
    else:
        # Try to derive from records' report_month or date_planted
        if records:
            max_date = None
            for rec in records:
                # Hybridization uses date_planted
                if module == 'hybridization':
                    d = getattr(rec, 'date_planted', None)
                    if d and (max_date is None or d > max_date):
                        max_date = d
                else:
                    d = getattr(rec, 'report_month', None)
                    if d and (max_date is None or d > max_date):
                        max_date = d
            if max_date:
                # For reports/months, use the last day of the month
                # For daily things (like hybridization), use the day itself
                if module == 'hybridization':
                    as_of_str = f'as of {max_date.strftime("%B %d, %Y")}'
                else:
                    last_day = calendar.monthrange(max_date.year, max_date.month)[1]
                    d_end = date(max_date.year, max_date.month, last_day)
                    as_of_str = f'as of {d_end.strftime("%B %d, %Y")}'
            else:
                # Default fallback
                today = date.today()
                first_of_month = today.replace(day=1)
                last_of_prev = first_of_month - timedelta(days=1)
                as_of_str = f'as of {last_of_prev.strftime("%B %d, %Y")}'
        else:
            today = date.today()
            first_of_month = today.replace(day=1)
            last_of_prev = first_of_month - timedelta(days=1)
            as_of_str = f'as of {last_of_prev.strftime("%B %d, %Y")}'

    for site_name, site_records in sites.items():
        ws = wb.create_sheet(title=site_name[:31])
        
        # Determine merge range based on module
        if module == 'harvest':
            merge_end = 'S'
        elif module == 'hybridization':
            merge_end = 'I'
        else:
            merge_end = 'I'

        ws.merge_cells(f"A1:{merge_end}1")
        ws.merge_cells(f"A2:{merge_end}2")
        ws.merge_cells(f"A3:{merge_end}3")
        ws.merge_cells(f"A4:{merge_end}4")
        
        ws["A1"] = "PHILIPPINE COCONUT AUTHORITY"
        ws["A2"] = "COCONUT HYBRIDIZATION PROJECT-CFIDP"
        ws["A3"] = "ON-FARM HYBRID SEEDNUT PRODUCTION" if module == 'harvest' else f"{module.replace('_', ' ').upper()} REPORT"
        ws["A4"] = as_of_str
        ws["A4"].font = Font(name='Calibri', size=10, underline='single')
        ws["A4"].alignment = Alignment(horizontal="center")
            
        for row in [1, 2, 3]:
            ws.cell(row=row, column=1).font = title_font
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

        _add_header_logo(ws)

        current_row = 7
        
        if module == 'harvest':
            headers = [
                ("A5", "A6", "Farm Location"),
                ("B5", "B6", "Name of Partner"),
                ("C5", "C6", "Area (Ha.)"),
                ("D5", "D6", "Age of Palms (Years)"),
                ("E5", "E6", "No. of Hybridized Palms"),
                ("F5", "F6", "Variety / Hybrid Crosses"),
                ("G5", "G6", "Seednuts Produced"),
                ("H5", "S5", "Monthly Production (No. of Seednuts)")
            ]
            
            for start, end, text in headers:
                ws.merge_cells(f"{start}:{end}")
                cell = ws[start]
                cell.value = text
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                
            months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            for i, month in enumerate(months):
                col = get_column_letter(8 + i)
                cell = ws[f"{col}6"]
                cell.value = month
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                ws.column_dimensions[col].width = 10

            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 30
            ws.column_dimensions['G'].width = 15

            for r in site_records:
                varieties = list(r.varieties.all()) if hasattr(r, 'varieties') else []
                variety_text = "\n".join([v.variety for v in varieties]) if varieties else getattr(r, 'variety', '')
                seednuts_text = "\n".join([str(v.seednuts_type) for v in varieties]) if varieties else getattr(r, 'seednuts_produced', '')
                
                row_data = [
                    r.location, r.farm_name, r.area_ha, r.age_of_palms, getattr(r, 'num_hybridized_palms', ''),
                    variety_text, seednuts_text,
                    r.production_jan, r.production_feb, r.production_mar,
                    r.production_apr, r.production_may, r.production_jun,
                    r.production_jul, r.production_aug, r.production_sep,
                    r.production_oct, r.production_nov, r.production_dec
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = value if value else ""
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
                    cell.border = thin_border
                current_row += 1
                
            _add_footer(ws, current_row, site_name, 19, records=site_records)

        elif module == 'hybridization':
            headers_list = ["Hybrid Code", "Crop Type", "Parent Line A", "Parent Line B",
                            "Date Planted", "Growth Status", "Status", "Notes", "Created By"]
            for col_idx, text in enumerate(headers_list, 1):
                col = get_column_letter(col_idx)
                cell = ws[f"{col}5"]
                cell.value = text
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                ws.column_dimensions[col].width = 18

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['H'].width = 35

            current_row = 6
            for r in site_records:
                row_data = [
                    r.hybrid_code, r.crop_type, r.parent_line_a, r.parent_line_b,
                    r.date_planted.strftime('%Y-%m-%d') if r.date_planted else '',
                    r.get_growth_status_display(), r.get_status_display(),
                    r.notes,
                    r.created_by.get_full_name() or r.created_by.username,
                ]
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = str(value) if value else ""
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
                    cell.border = thin_border
                current_row += 1
                
            _add_footer(ws, current_row, site_name, 9, records=site_records)

        # Apply Page Setup (Letter + Landscape)
        ws.page_setup.paperSize = 1  # 1 = Letter
        ws.page_setup.orientation = 'landscape'
        # Fit to one page wide
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0 
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        # Fallback basic format for other modules
        if module != 'harvest' and module != 'hybridization':
            headers_list = ["Record ID", "Report Month", "Site", "Created By"]
            for col_idx, text in enumerate(headers_list, 1):
                col = get_column_letter(col_idx)
                cell = ws[f"{col}5"]
                cell.value = text
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            
            current_row = 6
            for r in site_records:
                uname = r.created_by.get_full_name() if hasattr(r, 'created_by') else ""
                row_data = [
                    r.id, getattr(r, 'report_month', ''), 
                    r.field_site.name if r.field_site else "", uname
                ]
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = str(value)
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
                    cell.border = thin_border
                current_row += 1
                
            _add_footer(ws, current_row, site_name, 4, records=site_records)
        
        # Finally, auto-width for the whole sheet
        _excel_auto_width(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
