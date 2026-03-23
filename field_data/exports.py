"""
Excel export service for PCA field data.
Generates .xlsx files from database records matching the PCA report formats.
"""
import io
import calendar
from collections import OrderedDict
from datetime import date, timedelta
import os
from django.conf import settings
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.utils import column_index_from_string
from openpyxl.drawing.xdr import XDRPositiveSize2D


def _add_header_logo(ws):

    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'PCA&DA_Logo.png')

    if os.path.exists(logo_path):
        img = XLImage(logo_path)

        # Image size (pixels converted to EMU)
        width = 75
        height = 75

        img.width = width
        img.height = height

        col = column_index_from_string('B') - 1
        row = 0

        # Move logo to top left/center area without extreme offsets
        offset_x = 0   
        offset_y = 0   

        marker = AnchorMarker(col=col, colOff=offset_x,
                              row=row, rowOff=offset_y)

        size = XDRPositiveSize2D(width * 9525, height * 9525)

        img.anchor = OneCellAnchor(_from=marker, ext=size)

        ws.add_image(img)

def _add_footer(ws, start_row, site, num_cols, records=None):
    """Add the Prepared by, Reviewed by, and Noted by footer in a 3-column layout."""
    # Localize labels and signatories from FieldSite overrides
    prep_label = site.prepared_by_label if site and hasattr(site, 'prepared_by_label') and site.prepared_by_label else "Prepared by:"
    rev_label = site.reviewed_by_label if site and hasattr(site, 'reviewed_by_label') and site.reviewed_by_label else "Reviewed by:"
    note_label = site.noted_by_label if site and hasattr(site, 'noted_by_label') and site.noted_by_label else "Noted by:"

    ov_prep_name = site.prepared_by_name if site and hasattr(site, 'prepared_by_name') and site.prepared_by_name else ""
    ov_prep_title = site.prepared_by_title if site and hasattr(site, 'prepared_by_title') and site.prepared_by_title else ""
    ov_rev_name = site.reviewed_by_name if site and hasattr(site, 'reviewed_by_name') and site.reviewed_by_name else ""
    ov_rev_title = site.reviewed_by_title if site and hasattr(site, 'reviewed_by_title') and site.reviewed_by_title else ""
    ov_note_name = site.noted_by_name if site and hasattr(site, 'noted_by_name') and site.noted_by_name else ""
    ov_note_title = site.noted_by_title if site and hasattr(site, 'noted_by_title') and site.noted_by_title else ""

    prepared_user = None
    reviewed_user = None
    noted_user = None

    if records:
        for rec in reversed(list(records)):
            if not prepared_user and getattr(rec, 'prepared_by', None):
                prepared_user = rec.prepared_by
            if not reviewed_user and getattr(rec, 'reviewed_by', None):
                reviewed_user = rec.reviewed_by
            if not noted_user and getattr(rec, 'noted_by', None):
                noted_user = rec.noted_by

    # Final strings for Prepared By
    if ov_prep_name:
        prepared_by = ov_prep_name.upper()
        prepared_title = ov_prep_title
    elif prepared_user:
        mi = prepared_user.profile.middle_initial + ' ' if hasattr(prepared_user, 'profile') and prepared_user.profile.middle_initial else ''
        name_str = f"{prepared_user.first_name} {mi}{prepared_user.last_name}".strip().upper()
        prepared_by = name_str if name_str else prepared_user.username.upper()
        prepared_title = prepared_user.profile.get_role_display() if hasattr(prepared_user, 'profile') else 'COS/Agriculturist'
    else:
        prepared_by = '_______________________'
        prepared_title = 'COS/Agriculturist'

    # Reviewed By
    if ov_rev_name:
        reviewed_by = ov_rev_name.upper()
        reviewed_title = ov_rev_title
    elif reviewed_user:
        mi = reviewed_user.profile.middle_initial + ' ' if hasattr(reviewed_user, 'profile') and reviewed_user.profile.middle_initial else ''
        name_str = f"{reviewed_user.first_name} {mi}{reviewed_user.last_name}".strip().upper()
        reviewed_by = name_str if name_str else reviewed_user.username.upper()
        reviewed_title = reviewed_user.profile.get_role_display() if hasattr(reviewed_user, 'profile') else 'Senior Agriculturist'
    else:
        reviewed_by = '_______________________'
        reviewed_title = 'Senior Agriculturist'

    # Noted By
    if ov_note_name:
        noted_by = ov_note_name.upper()
        noted_title = ov_note_title
    elif noted_user:
        mi = noted_user.profile.middle_initial + ' ' if hasattr(noted_user, 'profile') and noted_user.profile.middle_initial else ''
        name_str = f"{noted_user.first_name} {mi}{noted_user.last_name}".strip().upper()
        noted_by = name_str if name_str else noted_user.username.upper()
        noted_title = noted_user.profile.get_role_display() if hasattr(noted_user, 'profile') else 'PCDM/Division Chief I'
    else:
        noted_by = '_______________________'
        noted_title = 'PCDM/Division Chief I'
    
    col1 = 1
    # Find a good middle column.
    col2 = max(2, (num_cols // 2) - 1)
    
    # Find a good right column that isn't cut off
    if num_cols >= 18:
        col3 = num_cols - 4
    elif num_cols >= 12:
        col3 = num_cols - 2
    else:
        col3 = max(3, num_cols - 1)
        
    row = start_row + 3
    
    center_font = Font(name='Calibri', size=10, bold=True)
    normal_font = Font(name='Calibri', size=10)
    center_align = Alignment(horizontal='center')
    left_align = Alignment(horizontal='left')

    # labels row
    ws.cell(row=row, column=col1, value=prep_label).font = center_font
    ws.cell(row=row, column=col1).alignment = left_align
    
    ws.cell(row=row, column=col2, value=rev_label).font = center_font
    ws.cell(row=row, column=col2).alignment = left_align

    ws.cell(row=row, column=col3, value=note_label).font = center_font
    ws.cell(row=row, column=col3).alignment = left_align
    
    row += 4
    
    # Names
    ws.cell(row=row, column=col1, value=prepared_by).font = center_font
    ws.cell(row=row, column=col1).alignment = center_align
    
    ws.cell(row=row, column=col2, value=reviewed_by).font = center_font
    ws.cell(row=row, column=col2).alignment = center_align
    
    ws.cell(row=row, column=col3, value=noted_by).font = center_font
    ws.cell(row=row, column=col3).alignment = center_align
    
    row += 1
    
    # Titles
    ws.cell(row=row, column=col1, value=prepared_title).font = normal_font
    ws.cell(row=row, column=col1).alignment = center_align
    
    ws.cell(row=row, column=col2, value=reviewed_title).font = normal_font
    ws.cell(row=row, column=col2).alignment = center_align
    
    ws.cell(row=row, column=col3, value=noted_title).font = normal_font
    ws.cell(row=row, column=col3).alignment = center_align
    
    # Insert signatures
    def insert_signature(user_obj, col_idx):
        if not user_obj or not hasattr(user_obj, 'profile'): return
        if not user_obj.profile.signature_image: return
        try:
            img_path = user_obj.profile.signature_image.path
            if os.path.exists(img_path):
                img = XLImage(img_path)
                # Max dimensions (approx 150x60)
                img.width = min(img.width, 150)
                img.height = min(img.height, 60)
                
                # Fetch col width to center it
                col_letter = get_column_letter(col_idx)
                col_width_chars = ws.column_dimensions[col_letter].width or 8.43
                col_width_pixels = col_width_chars * 7.5 # Approx pixels per char in Calibri 11
                
                # Centering calculation
                offset_pixels = max(0, (col_width_pixels - img.width) / 2)
                offset_emu = int(offset_pixels * 9525)
                
                # anchor it above the name
                marker = AnchorMarker(col=col_idx-1, colOff=offset_emu, row=row-3, rowOff=0)
                size = XDRPositiveSize2D(int(img.width * 9525), int(img.height * 9525))
                img.anchor = OneCellAnchor(_from=marker, ext=size)
                ws.add_image(img)
        except Exception:
            pass

    insert_signature(prepared_user if not ov_prep_name else None, col1)
    insert_signature(reviewed_user if not ov_rev_name else None, col2)
    insert_signature(noted_user if not ov_note_name else None, col3)

    return row



def _build_as_of_str(as_of_date=None, records=None):
    """Build the 'as of ...' date string.

    Priority: explicit as_of_date > max report_month from records > last day of prev month.
    """
    if as_of_date:
        return f'as of {as_of_date.strftime("%B %d, %Y")}'
    # Try to derive from records' report_month
    if records:
        max_month = None
        for rec in records:
            rm = getattr(rec, 'report_month', None)
            if rm and (max_month is None or rm > max_month):
                max_month = rm
        if max_month:
            last_day = calendar.monthrange(max_month.year, max_month.month)[1]
            d = date(max_month.year, max_month.month, last_day)
            return f'as of {d.strftime("%B %d, %Y")}'
    # Fallback
    today = date.today()
    first_of_month = today.replace(day=1)
    last_of_prev = first_of_month - timedelta(days=1)
    return f'as of {last_of_prev.strftime("%B %d, %Y")}'


def _style_header(ws, row_num, num_cols):
    """Apply PCA green header style to a row."""
    green_fill = PatternFill(start_color='0B9E4F', end_color='0B9E4F', fill_type='solid')
    white_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = green_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border


def _apply_data_border(ws, row_num, num_cols):
    """Apply thin borders and standard alignment to data cells."""
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)


def _auto_width(ws, num_cols, skip_row=1):
    """Auto-size column widths with a minimum buffer for editing."""
    for col_idx in range(1, num_cols + 1):
        max_len = 0
        column_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=skip_row):
            for cell in row:
                try:
                    if not isinstance(cell, MergedCell) and cell.value:
                        # Handle multi-line text
                        lines = str(cell.value).split('\n')
                        line_max = max([len(l) for l in lines])
                        if line_max > max_len:
                            max_len = line_max
                except:
                    pass
        # Min width of 10 for better fit on Letter paper, max of 40
        ws.column_dimensions[column_letter].width = max(10, min(max_len + 2, 40))


def _title_rows(ws, title, subtitle, site_name, merge_end='S', as_of_date=None, records=None):
    """Write PCA header rows and configure page setup."""
    # Page Setup (Letter + Landscape)
    ws.page_setup.paperSize = 1  # 1 = Letter
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.merge_cells(f'A1:{merge_end}1')
    ws['A1'].value = 'Department of Agriculture'
    ws['A1'].font = Font(name='Calibri', size=10)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A2:{merge_end}2')
    ws['A2'].value = 'PHILIPPINE COCONUT AUTHORITY'
    ws['A2'].font = Font(name='Calibri', size=11, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A3:{merge_end}3')
    ws['A3'].value = 'REGION VII'
    ws['A3'].font = Font(name='Calibri', size=10)
    ws['A3'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A4:{merge_end}4')
    ws['A4'].value = _build_as_of_str(as_of_date, records)
    ws['A4'].font = Font(name='Calibri', size=10, underline='single')
    ws['A4'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A5:{merge_end}5')
    ws['A5'].value = title
    ws['A5'].font = Font(name='Calibri', size=11, bold=True)
    ws['A5'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A6:{merge_end}6')
    ws['A6'].value = subtitle
    ws['A6'].font = Font(name='Calibri', size=10)
    ws['A6'].alignment = Alignment(horizontal='center')


# ---------------------------------------------------------------------------
# Export: Hybrid Distribution
# ---------------------------------------------------------------------------

def export_distribution(queryset, field_site_name='All Sites', as_of_date=None):
    """Export Hybrid Distribution records matching PCA format, separated by farm."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sites = {}
    for rec in queryset:
        s_name = rec.field_site.name if rec.field_site else 'Unknown Site'
        if s_name not in sites:
            sites[s_name] = []
        sites[s_name].append(rec)

    if not sites:
        ws = wb.create_sheet('No Data')
        ws.append(['No records found.'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    for site_name, site_records in sites.items():
        ws = wb.create_sheet(title=site_name[:31])

        _title_rows(ws,
            'COCONUT HYBRIDIZATION PROGRAM',
            'COMMUNAL NURSERY: DISPATCHED SEEDLINGS',
            site_name, merge_end='S', as_of_date=as_of_date, records=site_records)
        _add_header_logo(ws)

        # Headers - Row 8 (main), Row 9 (sub-headers)
        main_headers = {
            'A': 'Region', 'B': 'Province', 'C': 'District', 'D': 'Municipality',
            'E': 'Barangay', 'F': 'Name of Farmer Participant', 'I': 'Gender',
            'K': 'Farm Location', 'N': 'Seedlings Received',
            'O': 'Date Received', 'P': 'Type/Variety',
            'Q': 'No. of Seedlings Planted', 'R': 'Date Planted', 'S': 'REMARKS',
        }
        sub_headers_9 = {'F': 'Family Name', 'G': 'Given Name', 'H': 'M.I.'}
        sub_headers_10 = {'I': 'Male', 'J': 'Female', 'K': 'Barangay', 'L': 'Municipality', 'M': 'Province'}

        for col_letter, val in main_headers.items():
            ws[f'{col_letter}8'] = val
        for col_letter, val in sub_headers_9.items():
            ws[f'{col_letter}9'] = val
        for col_letter, val in sub_headers_10.items():
            ws[f'{col_letter}10'] = val

        _style_header(ws, 8, 19)
        _style_header(ws, 9, 19)
        _style_header(ws, 10, 19)

        # Custom Insertions below headers
        ws.merge_cells('A11:S11')
        ws['A11'].value = 'BOHOL PROVINCE'
        ws['A11'].font = Font(name='Calibri', size=11, bold=True)
        ws['A11'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A12:S12')
        ws['A12'].value = f'COMMUNAL NURSERY AT {site_name}'
        ws['A12'].font = Font(name='Calibri', size=11, bold=True)
        ws['A12'].alignment = Alignment(horizontal='center')

        # Data rows start at 13
        row = 13
        total_planted = 0
        total_received = 0
        
        for rec in site_records:
            ws.cell(row=row, column=1, value=rec.region)
            ws.cell(row=row, column=2, value=rec.province)
            ws.cell(row=row, column=3, value=rec.district)
            ws.cell(row=row, column=4, value=rec.municipality)
            ws.cell(row=row, column=5, value=rec.barangay)
            ws.cell(row=row, column=6, value=rec.farmer_last_name)
            ws.cell(row=row, column=7, value=rec.farmer_first_name)
            ws.cell(row=row, column=8, value=rec.farmer_middle_initial)
            ws.cell(row=row, column=9, value='/' if rec.is_male else '')
            ws.cell(row=row, column=10, value='/' if rec.is_female else '')
            ws.cell(row=row, column=11, value=rec.farm_barangay)
            ws.cell(row=row, column=12, value=rec.farm_municipality)
            ws.cell(row=row, column=13, value=rec.farm_province)
            ws.cell(row=row, column=14, value=rec.seedlings_received)
            ws.cell(row=row, column=15, value=rec.date_received.strftime('%m/%d/%Y') if rec.date_received else '')
            ws.cell(row=row, column=16, value=rec.variety)
            ws.cell(row=row, column=17, value=rec.seedlings_planted)
            ws.cell(row=row, column=18, value=rec.date_planted.strftime('%m/%d/%Y') if rec.date_planted else '')
            ws.cell(row=row, column=19, value=rec.remarks)
            _apply_data_border(ws, row, 19)
            row += 1
            
            # Tally for footer
            try:
                total_planted += int(rec.seedlings_planted)
            except (ValueError, TypeError):
                pass
            try:
                val = str(rec.seedlings_received).replace(',', '')
                total_received += int(val)
            except (ValueError, TypeError, AttributeError):
                pass
                
        # Total Row
        ws.cell(row=row, column=6, value="TOTAL:")
        ws.cell(row=row, column=6).font = Font(name='Calibri', size=11, bold=True)
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
        
        if total_received > 0:
            ws.cell(row=row, column=14, value=total_received)
            ws.cell(row=row, column=14).font = Font(name='Calibri', size=11, bold=True)
            
        ws.cell(row=row, column=17, value=total_planted)
        ws.cell(row=row, column=17).font = Font(name='Calibri', size=11, bold=True)
        
        for c in range(1, 20):
            ws.cell(row=row, column=c).border = Border(
                top=Side(style='thick'), bottom=Side(style='thin'),
                left=Side(style='thin'), right=Side(style='thin')
            )
        row += 2

        _auto_width(ws, 19)
        site_obj = site_records[0].field_site if site_records else None
        _add_footer(ws, row, site_obj, 19, records=site_records)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Export: Monthly Harvest
# ---------------------------------------------------------------------------

def export_harvest(queryset, field_site_name='All Sites', as_of_date=None):
    """Export Monthly Harvest records matching PCA format, separated by farm."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Group by FieldSite name
    sites = {}
    for rec in queryset.prefetch_related('varieties'):
        s_name = rec.field_site.name if rec.field_site else 'Unknown Site'
        if s_name not in sites:
            sites[s_name] = []
        sites[s_name].append(rec)

    if not sites:
        ws = wb.create_sheet('No Data')
        ws.append(['No records found.'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    for site_name, site_records in sites.items():
        ws = wb.create_sheet(title=site_name[:31])

        # Apply Page Setup (Letter + Landscape)
        ws.page_setup.paperSize = 1  # 1 = Letter
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        # Title rows
        ws.merge_cells('A1:U1')
        ws['A1'].value = 'PHILIPPINE COCONUT AUTHORITY'
        ws['A1'].font = Font(name='Calibri', size=11, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:U2')
        ws['A2'].value = 'COCONUT HYBRIDIZATION PROJECT-CFIDP'
        ws['A2'].font = Font(name='Calibri', size=10)
        ws['A2'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A3:U3')
        ws['A3'].value = 'ON-FARM HYBRID SEEDNUT PRODUCTION'
        ws['A3'].font = Font(name='Calibri', size=10)
        ws['A3'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A4:U4')
        ws['A4'].value = _build_as_of_str(as_of_date, site_records)
        ws['A4'].font = Font(name='Calibri', size=10, underline='single')
        ws['A4'].alignment = Alignment(horizontal='center')

        _add_header_logo(ws)

        # Headers
        main_h = {
            'A': 'Farm Location', 'B': 'Name of Partner', 'C': 'Area (Ha.)',
            'D': 'Age of Palms (Years)', 'E': 'No. of Hybridized Palms',
            'F': 'Variety / Hybrid Crosses', 'G': 'Seednuts Produced',
            'H': 'Monthly Production (No. of Seednuts)',
            'T': 'TOTAL', 'U': 'Remarks'
        }
        month_h = {
            'H': 'Jan', 'I': 'Feb', 'J': 'Mar', 'K': 'Apr', 'L': 'May', 'M': 'Jun',
            'N': 'Jul', 'O': 'Aug', 'P': 'Sep', 'Q': 'Oct', 'R': 'Nov', 'S': 'Dec',
        }

        for col, val in main_h.items():
            ws[f'{col}5'] = val
        for col, val in month_h.items():
            ws[f'{col}6'] = val
        _style_header(ws, 5, 21)
        _style_header(ws, 6, 21)

        # Merge the two extra header columns vertically since they don't have sub-headers
        ws.merge_cells('T5:T6')
        ws.merge_cells('U5:U6')

        row = 7
        # Track column totals for final TOTAL row
        month_totals = {m: 0 for m in range(8, 20)}  # cols 8-19 = Jan-Dec
        grand_total = 0
        area_total = 0
        palms_total = 0

        # ---- Group records by farm (location + farm_name) ----
        farms = OrderedDict()
        for rec in site_records:
            key = (rec.location or '', rec.farm_name or '')
            if key not in farms:
                farms[key] = {
                    'location': rec.location,
                    'farm_name': rec.farm_name,
                    'area_ha': rec.area_ha,
                    'age_of_palms': rec.age_of_palms,
                    'num_hybridized_palms': rec.num_hybridized_palms,
                    # variety_key -> {seednuts_type, month_counts{col: count}, remarks}
                    'varieties': OrderedDict(),
                }
            farm = farms[key]
            month_col = 7 + rec.report_month.month  # month 1(Jan)=col 8, etc.
            
            # If exporting for a specific year, skip counts from other years (e.g. from the carry forward logic)
            # Actually, `site_records` is already filtered by year matching in the view (`_apply_date_filters`),
            # so rec.report_month is already guaranteed to be in the requested year.
            
            for var in rec.varieties.all():
                var_key = (var.variety or '', var.seednuts_type or '')
                if var_key not in farm['varieties']:
                    farm['varieties'][var_key] = {
                        'variety': var.variety,
                        'seednuts_type': var.seednuts_type,
                        'month_counts': {},
                        'remarks': var.remarks or '',
                    }
                vdata = farm['varieties'][var_key]
                count = var.seednuts_count or 0
                if count:
                    vdata['month_counts'][month_col] = vdata['month_counts'].get(month_col, 0) + count

        # ---- Write rows per farm ----
        for (loc, fname), farm in farms.items():
            first_row = True
            try:
                area_total += float(farm['area_ha']) if farm['area_ha'] else 0
            except (ValueError, TypeError):
                pass
            palms_total += farm['num_hybridized_palms'] or 0

            for var_key, vdata in farm['varieties'].items():
                if first_row:
                    ws.cell(row=row, column=1, value=farm['location'])
                    ws.cell(row=row, column=2, value=farm['farm_name'])
                    ws.cell(row=row, column=3, value=farm['area_ha'])
                    ws.cell(row=row, column=4, value=farm['age_of_palms'])
                    ws.cell(row=row, column=5, value=farm['num_hybridized_palms'])
                    first_row = False

                ws.cell(row=row, column=6, value=vdata['variety'])
                ws.cell(row=row, column=7, value=vdata['seednuts_type'])

                # Place seednuts in correct month columns
                var_total = 0
                for mcol, mcount in vdata['month_counts'].items():
                    ws.cell(row=row, column=mcol, value=mcount)
                    month_totals[mcol] = month_totals.get(mcol, 0) + mcount
                    var_total += mcount
                    grand_total += mcount

                if var_total:
                    ws.cell(row=row, column=20, value=var_total)  # TOTAL col

                ws.cell(row=row, column=21, value=vdata['remarks'])
                _apply_data_border(ws, row, 21)
                row += 1

        # Grand TOTAL row
        ws.cell(row=row, column=2, value='TOTAL')
        ws.cell(row=row, column=3, value=area_total)
        ws.cell(row=row, column=5, value=palms_total)
        for col_idx, total in month_totals.items():
            if total:
                ws.cell(row=row, column=col_idx, value=total)
        if grand_total:
            ws.cell(row=row, column=20, value=grand_total)
        _apply_data_border(ws, row, 21)
        # Bold the total row
        for c in range(1, 22):
            cell = ws.cell(row=row, column=c)
            cell.font = Font(name='Calibri', size=10, bold=True)

        _auto_width(ws, 21)
        site_obj = site_records[0].field_site if site_records else None
        _add_footer(ws, row, site_obj, 21, records=site_records)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Export: Nursery Operations
# ---------------------------------------------------------------------------

def export_nursery(queryset, field_site_name='All Sites', as_of_date=None):
    """Export Nursery Operations records matching PCA format, separated by farm."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sites = {}
    for rec in queryset.prefetch_related('batches'):
        s_name = rec.field_site.name if rec.field_site else 'Unknown Site'
        if s_name not in sites:
            sites[s_name] = []
        sites[s_name].append(rec)

    if not sites:
        ws = wb.create_sheet('No Data')
        ws.append(['No records found.'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    for site_name, site_records in sites.items():
        ws = wb.create_sheet(title=site_name[:31])

        # Apply Page Setup (Letter + Landscape)
        ws.page_setup.paperSize = 1  # 1 = Letter
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        # Title
        ws.merge_cells('A1:R1')
        ws['A1'].value = 'COCONUT HYBRIDIZATION PROGRAM-CFIDP'
        ws['A1'].font = Font(name='Calibri', size=11, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:R2')
        ws['A2'].value = 'COMMUNAL NURSERY ESTABLISHMENT'
        ws['A2'].font = Font(name='Calibri', size=10)
        ws['A2'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A3:R3')
        ws['A3'].value = 'Communal Nursery Establishment Monthly Report'
        ws['A3'].font = Font(name='Calibri', size=10)
        ws['A3'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A4:R4')
        ws['A4'].value = _build_as_of_str(as_of_date, site_records)
        ws['A4'].font = Font(name='Calibri', size=10, underline='single')
        ws['A4'].alignment = Alignment(horizontal='center')

        _add_header_logo(ws)

        # Headers
        headers = [
            'Region / Province / District',  # A
            'Barangay / Municipality',        # B
            'Entity Name',                     # C
            'Representative',                 # D
            'Target No. of Seednuts',         # E
            'No. of Seednuts Harvested',      # F
            'Date Harvested',                 # G
            'Date Seednuts Received',         # H
            'Source of Seednuts',             # I
            'Type/Variety',                   # J
            'No. of Seednuts Sown',          # K
            'Date Seednut Sown',             # L
            'No. of Seedlings Germinated',   # M
            'No. of Ungerminated Seednuts',  # N
            'No. of Culled Seedlings',       # O
            'No. of Good Seedlings @ 1 ft',  # P
            'No. of Ready to Plant (Polybagged)',  # Q
            'No. of Seedlings Dispatched',   # R
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=5, column=col, value=h)
        _style_header(ws, 5, len(headers))

        row = 6
        for rec in site_records:
            batches = rec.batches.all()
            if not batches:
                # Still output an empty row for the parent if it has no batches
                vals = [
                    rec.region_province_district,
                    rec.barangay_municipality,
                    rec.proponent_entity,
                    rec.proponent_representative,
                    rec.target_seednuts,
                    '', '', '', '', '', '', '', '', '', '', '', '', ''
                ]
                for col, v in enumerate(vals, 1):
                    ws.cell(row=row, column=col, value=v)
                _apply_data_border(ws, row, len(headers))
                row += 1
                continue

            start_row = row
            for batch in batches:
                vals = [
                    rec.region_province_district,
                    rec.barangay_municipality,
                    rec.proponent_entity,
                    rec.proponent_representative,
                    rec.target_seednuts,
                    batch.seednuts_harvested,
                    batch.date_harvested,
                    batch.date_received,
                    batch.source_of_seednuts,
                    batch.variety,
                    batch.seednuts_sown,
                    batch.date_sown,
                    batch.seedlings_germinated,
                    batch.ungerminated_seednuts,
                    batch.culled_seedlings,
                    batch.good_seedlings,
                    batch.ready_to_plant,
                    batch.seedlings_dispatched,
                ]
                for col, v in enumerate(vals, 1):
                    ws.cell(row=row, column=col, value=v)
                _apply_data_border(ws, row, len(headers))
                row += 1
            
            # Merge parent columns if there are multiple batches
            end_row = row - 1
            if end_row > start_row:
                for col_idx in range(1, 6): # Columns A to E (1 to 5)
                    ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
                    cell = ws.cell(row=start_row, column=col_idx)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        _auto_width(ws, len(headers))
        _add_footer(ws, row, site_name, len(headers), records=site_records)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Export: Pollen Production
# ---------------------------------------------------------------------------

def export_pollen(queryset, field_site_name='All Sites', as_of_date=None):
    """Export Pollen Production records matching PCA format, separated by farm."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sites = {}
    for rec in queryset:
        s_name = rec.field_site.name if rec.field_site else 'Unknown Site'
        if s_name not in sites:
            sites[s_name] = []
        sites[s_name].append(rec)

    if not sites:
        ws = wb.create_sheet('No Data')
        ws.append(['No records found.'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    for site_name, site_records in sites.items():
        ws = wb.create_sheet(title=site_name[:31])

        # Apply Page Setup (Letter + Landscape)
        ws.page_setup.paperSize = 1  # 1 = Letter
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        # Title
        ws.merge_cells('A1:L1')
        ws['A1'].value = 'COCONUT HYBRIDIZATION PROGRAM-CFIDP'
        ws['A1'].font = Font(name='Calibri', size=11, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:L2')
        ws['A2'].value = 'POLLEN PRODUCTION'
        ws['A2'].font = Font(name='Calibri', size=10)
        ws['A2'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A3:L3')
        ws['A3'].value = f'Pollen Production and Inventory Monthly Report — {site_name}'
        ws['A3'].font = Font(name='Calibri', size=10)
        ws['A3'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A4:L4')
        ws['A4'].value = _build_as_of_str(as_of_date, site_records)
        ws['A4'].font = Font(name='Calibri', size=10, underline='single')
        ws['A4'].alignment = Alignment(horizontal='center')

        _add_header_logo(ws)

        # Determine Center/Unit text based on site name
        center_text = 'Unknown'
        if 'Balilihan' in site_name:
            center_text = 'ON FARM BALILIHAN'
        elif 'Loay' in site_name:
            center_text = 'LOAY CODE FARM, LAS SALINAS SUR, LOAY, BOHOL'
        
        ws.merge_cells('A6:L6')
        ws['A6'].value = f'CENTER/UNIT: {center_text}'
        ws['A6'].font = Font(name='Calibri', size=11, bold=True)
        ws['A6'].alignment = Alignment(horizontal='left')
        
        # Get pollen variety from first record
        pollen_variety_text = site_records[0].pollen_variety if site_records and site_records[0].pollen_variety else ''
        ws.merge_cells('A7:L7')
        ws['A7'].value = f'POLLEN VARIETY: {pollen_variety_text}'
        ws['A7'].font = Font(name='Calibri', size=11, bold=True)
        ws['A7'].alignment = Alignment(horizontal='left')

        # Headers (row 9 main, row 10 sub)
        main_h = {
            'A': 'MONTH',
            'B': 'Ending Balance\nLast Month\n(g Pollens)',
            'C': 'POLLENS RECEIVED FROM OTHER CENTER',
            'F': 'POLLEN UTILIZATION (grams of Pollen) per Week',
            'L': 'Ending Balance\n(g Pollens)',
        }
        sub_h = {
            'C': 'Source', 'D': 'Date Received\nmm/dd/yyyy', 'E': 'Grams of\nPollens',
            'F': 'Week 1', 'G': 'Week 2', 'H': 'Week 3',
            'I': 'Week 4', 'J': 'Week 5', 'K': 'TOTAL',
        }
        for col, val in main_h.items():
            cell = ws[f'{col}9']
            cell.value = val
            ws.merge_cells(f'{col}9:{col}10') if col in ['A', 'B', 'L'] else None
        
        # specific merges for grouped headers
        ws.merge_cells('C9:E9')
        ws.merge_cells('F9:K9')

        for col, val in sub_h.items():
            ws[f'{col}10'] = val
            
        _style_header(ws, 9, 12)
        _style_header(ws, 10, 12)

        import re
        def parse_num(val):
            if not val: return 0
            cleaned = re.sub(r'[^\d.-]', '', str(val))
            try:
                return float(cleaned)
            except ValueError:
                return 0

        row = 11
        total_received = 0
        total_util = 0

        for rec in site_records:
            ws.cell(row=row, column=1, value=rec.month_label)
            ws.cell(row=row, column=2, value=f"{parse_num(rec.ending_balance_prev):g} g" if parse_num(rec.ending_balance_prev) else "")
            ws.cell(row=row, column=3, value=rec.pollen_source)
            ws.cell(row=row, column=4, value=rec.date_received)
            ws.cell(row=row, column=5, value=f"{parse_num(rec.pollens_received):g} g" if parse_num(rec.pollens_received) else "")
            ws.cell(row=row, column=6, value=f"{parse_num(rec.week1):g} g" if parse_num(rec.week1) else "")
            ws.cell(row=row, column=7, value=f"{parse_num(rec.week2):g} g" if parse_num(rec.week2) else "")
            ws.cell(row=row, column=8, value=f"{parse_num(rec.week3):g} g" if parse_num(rec.week3) else "")
            ws.cell(row=row, column=9, value=f"{parse_num(rec.week4):g} g" if parse_num(rec.week4) else "")
            ws.cell(row=row, column=10, value=f"{parse_num(rec.week5):g} g" if parse_num(rec.week5) else "")
            ws.cell(row=row, column=11, value=f"{parse_num(rec.total_utilization):g} g" if parse_num(rec.total_utilization) else "")
            ws.cell(row=row, column=12, value=f"{parse_num(rec.ending_balance):g} g" if parse_num(rec.ending_balance) else "")
            _apply_data_border(ws, row, 12)
            
            total_received += parse_num(rec.pollens_received)
            total_util += parse_num(rec.total_utilization)
            
            row += 1

        # Total Row
        ws.cell(row=row, column=1, value="TOTAL:")
        ws.cell(row=row, column=1).font = Font(name='Calibri', size=11, bold=True)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='right')
        
        def format_weight(grams):
            if grams >= 1000:
                return f"{grams / 1000:,.2f} kg"
            return f"{grams:,.2f} g"
            
        if total_received > 0:
            ws.cell(row=row, column=5, value=format_weight(total_received))
            ws.cell(row=row, column=5).font = Font(name='Calibri', size=11, bold=True)
            
        if total_util > 0:
            ws.cell(row=row, column=11, value=format_weight(total_util))
            ws.cell(row=row, column=11).font = Font(name='Calibri', size=11, bold=True)

        for c in range(1, 13):
            ws.cell(row=row, column=c).border = Border(
                top=Side(style='thick'), bottom=Side(style='thin'),
                left=Side(style='thin'), right=Side(style='thin')
            )
        row += 2

        _auto_width(ws, 12)
        site_obj = site_records[0].field_site if site_records else None
        _add_footer(ws, row, site_obj, 12, records=site_records)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_terminal(queryset, field_site_name='All Sites', as_of_date=None):
    """Export Terminal Report records, separated by farm."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sites = {}
    for rec in queryset:
        s_name = rec.field_site.name if rec.field_site else 'Unknown Site'
        if s_name not in sites:
            sites[s_name] = []
        sites[s_name].append(rec)

    if not sites:
        ws = wb.create_sheet('No Data')
        ws.append(['No records found.'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    for site_name, site_records in sites.items():
        ws = wb.create_sheet(title=site_name[:31])

        # Apply Page Setup (Letter + Landscape)
        ws.page_setup.paperSize = 1  # 1 = Letter
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        ws.merge_cells('A1:R1')
        ws['A1'].value = 'COCONUT HYBRIDIZATION PROGRAM-CFIDP'
        ws['A1'].font = Font(name='Calibri', size=11, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:R2')
        ws['A2'].value = 'COMMUNAL NURSERY ESTABLISHMENT'
        ws['A2'].font = Font(name='Calibri', size=10)
        ws['A2'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A3:R3')
        ws['A3'].value = 'Communal Nursery Establishment Terminal Report'
        ws['A3'].font = Font(name='Calibri', size=10)
        ws['A3'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A4:R4')
        ws['A4'].value = _build_as_of_str(as_of_date, site_records)
        ws['A4'].font = Font(name='Calibri', size=10, underline='single')
        ws['A4'].alignment = Alignment(horizontal='center')
        
        _add_header_logo(ws)

        headers = [
            'Region / Province / District', 'Barangay / Municipality',
            'Entity Name', 'Representative', 'Target No. of Seednuts',
            'No. of Seednuts Harvested', 'Date Harvested',
            'Date Seednuts Received', 'Source of Seednuts', 'Type/Variety',
            'No. of Seednuts Sown', 'Date Seednut Sown',
            'No. of Seedlings Germinated', 'No. of Ungerminated Seednuts',
            'No. of Culled Seedlings', 'No. of Good Seedlings @ 1 ft',
            'No. of Ready to Plant (Polybagged)', 'No. of Seedlings Dispatched',
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=5, column=col, value=h)
        _style_header(ws, 5, len(headers))

        row = 6
        for rec in site_records:
            vals = [
                rec.region_province_district, rec.barangay_municipality,
                rec.proponent_entity, rec.proponent_representative,
                rec.target_seednuts, rec.seednuts_harvested, rec.date_harvested,
                rec.date_received, rec.source_of_seednuts, rec.variety,
                rec.seednuts_sown, rec.date_sown, rec.seedlings_germinated,
                rec.ungerminated_seednuts, rec.culled_seedlings, rec.good_seedlings,
                rec.ready_to_plant, rec.seedlings_dispatched,
            ]
            for col, v in enumerate(vals, 1):
                ws.cell(row=row, column=col, value=v)
            _apply_data_border(ws, row, len(headers))
            row += 1

        _auto_width(ws, len(headers))
        site_obj = site_records[0].field_site if site_records else None
        _add_footer(ws, row, site_obj, len(headers), records=site_records)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# Dispatcher
EXPORTERS = {
    'distribution': export_distribution,
    'harvest': export_harvest,
    'nursery': export_nursery,
    'terminal': export_terminal,
    'pollen': export_pollen,
}
